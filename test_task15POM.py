from datetime import datetime

import openpyxl


from task15_DDT.POM.login import orange_login

def read_excel(file_path,sheet_name='data'):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    rows = []
    for row in sheet.iter_rows(min_row=2,values_only=True):#skip header
        rows.append(list(row)) #converts tuple to list
    return rows,workbook,sheet

def write_test_result(sheet,row,result,tester='meenakshi'):
    sheet.cell(row=row, column=7).value = result
    sheet.cell(row=row, column=6).value = tester
    sheet.cell(row=row, column=4).value = datetime.today().strftime("%Y-%m-%d")
    sheet.cell(row=row, column=5).value = datetime.now().strftime("%H:%M:%S")

file_path = "C:/Users/HP/PycharmProjects/pythonProject/task15_DDT/data.xlsx"

def test_login(setup):
    try:
        driver = setup

        #invoking class
        login_page = orange_login(driver)

        #write back result to excel
        rows,workbook,sheet = read_excel(file_path)

        for index,row in enumerate(rows, start=2):
            test_id,username,password,Date,TimeofTest,tester,result = row
            login_page.login(username, password)

            if login_page.is_login_success():
                write_test_result(sheet,index,"Passed")
            else:
                write_test_result(sheet,index,"Failed")
            workbook.save(file_path)

    except Exception as e:
        print(e)


